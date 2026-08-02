# Экспорт учёта в Excel: операции за период + текущие долги, остатки, кассы.
import io
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import db
import prices

OP_TYPES = {
    "invoice": "Накладная",
    "payment": "Оплата",
    "transfer": "Перемещение",
    "return": "Возврат",
    "inventory": "Инвентаризация",
    "handover": "Инкассация",
    "writeoff": "Списание",
}

HEADER_FILL = PatternFill("solid", fgColor="1B5E20")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _sheet_header(ws, headers, widths):
    ws.append(headers)
    for col, width in enumerate(widths, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def build_export(start_iso: str, period_label: str) -> io.BytesIO:
    wb = Workbook()

    # --- Операции ---
    ws = wb.active
    ws.title = "Операции"
    _sheet_header(
        ws,
        ["№", "Дата", "Тип", "Сотрудник", "Склад", "Клиент",
         "Сумма, сом", "Оплата, сом", "Статус", "Описание"],
        [7, 17, 15, 14, 12, 18, 12, 12, 11, 45],
    )
    for op in db.operations_all_since(start_iso):
        try:
            data = json.loads(op["data"])
        except (ValueError, TypeError):
            data = {}
        try:
            dt = datetime.fromisoformat(op["ts"]).strftime("%d.%m.%Y %H:%M")
        except ValueError:
            dt = op["ts"]
        user = db.get_user(op["user_id"])
        wh = db.warehouse_by_id(op["warehouse_id"]) if op["warehouse_id"] else None
        client = db.client_get(op["client_id"]) if op["client_id"] else None
        total = ""
        payment = ""
        # У отменённых операций суммы не заполняем: сумма по колонке в
        # Excel должна сходиться с учётом, а сторнированные суммы в ней
        # завышали итог (сумма видна в «Описании», статус — «отменена»).
        if op["status"] == "done":
            if op["type"] == "invoice":
                total = data.get("total", "")
                payment = data.get("payment", "") or ""
            elif op["type"] in ("payment", "handover"):
                payment = data.get("amount", "")
            elif op["type"] in ("return", "writeoff"):
                total = -data.get("total", 0)
        ws.append([
            op["id"], dt, OP_TYPES.get(op["type"], op["type"]),
            user["name"] if user else op["user_id"],
            wh["name"] if wh else "",
            client["name"] if client else "",
            total, payment,
            "проведена" if op["status"] == "done" else "отменена",
            op["summary"],
        ])

    # --- Долги ---
    ws = wb.create_sheet("Долги")
    _sheet_header(ws, ["Склад", "Клиент", "Долг, сом"], [14, 24, 14])
    total_debt = 0.0
    for wh in db.all_warehouses():
        for c in db.clients_of(wh["id"]):
            if c["debt"]:
                ws.append([wh["name"], c["name"], c["debt"]])
                total_debt += c["debt"]
    ws.append([])
    ws.append(["", "ИТОГО", total_debt])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=3).font = Font(bold=True)

    # --- Остатки ---
    ws = wb.create_sheet("Остатки")
    _sheet_header(
        ws,
        ["Склад", "Товар", "Фасовка", "Кол-во, шт", "Цена, сом", "Сумма, сом"],
        [14, 45, 14, 12, 12, 14],
    )
    grand = 0.0
    for wh in db.all_warehouses():
        smap = db.stock_map(wh["id"])
        for p in prices.PRICE_LIST_DATA:
            qty = smap.get(p["id"], 0)
            if qty:
                amount = qty * p["price"]
                grand += amount
                ws.append([wh["name"], p["name"], p["volume"], qty, p["price"], amount])
        # Товар, выведенный из прайса, но с остатком: раньше молча выпадал
        # из экспорта, и остатки «не бились» с /stock.
        known = {p["id"] for p in prices.PRICE_LIST_DATA}
        for pid, qty in sorted(smap.items()):
            if qty and pid not in known:
                row = db.connect().execute(
                    "SELECT name, volume, price FROM products WHERE id=?",
                    (pid,)).fetchone()
                name = (row["name"] + " (вне прайса)") if row else f"товар №{pid} (вне прайса)"
                price = row["price"] if row else 0
                amount = qty * price
                grand += amount
                ws.append([wh["name"], name, row["volume"] if row else "",
                           qty, price, amount])
    ws.append([])
    ws.append(["", "ИТОГО по прайсу", "", "", "", grand])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)

    # --- Кассы ---
    ws = wb.create_sheet("Кассы")
    _sheet_header(ws, ["Сотрудник", "Наличные на руках, сом"], [20, 24])
    total_cash = 0.0
    # Включая уволенных: несданная касса бывшего сотрудника — всё ещё
    # деньги компании, раньше она молча пропадала из экспорта.
    for u in db.list_users(active_only=False):
        cash = db.cash_on_hand(u["id"])
        if cash:
            name = u["name"] if u["active"] else f"{u['name']} (уволен)"
            ws.append([name, cash])
            total_cash += cash
    ws.append([])
    ws.append(["ИТОГО", total_cash])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    # --- Инфо ---
    ws = wb.create_sheet("Инфо")
    ws.append(["Экспорт ВЕТОП"])
    ws.append([f"Период операций: {period_label}"])
    ws.append([f"Сформирован: {datetime.now(db.BISHKEK).strftime('%d.%m.%Y %H:%M')}"])
    ws.append(["Долги, остатки и кассы — на момент выгрузки."])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
